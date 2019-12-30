import ssl
from openpyxl import load_workbook

# 打开文件
workbook = load_workbook('e.xlsx')
# 获取sheet
book_sheet = workbook.active
print(book_sheet)
# 创建数组
temp_list = list()

# 获取sheet页的行数
row = book_sheet.max_row

# 迭代所有的行 从第2行开始
for row in range(2, row+1):
    # 依次获取5列数据
    cell_data_1 = book_sheet.cell(row, column=1).value
    cell_data_2 = book_sheet.cell(row, column=2).value
    cell_data_3 = book_sheet.cell(row, column=3).value
    cell_data_4 = book_sheet.cell(row, column=4).value
    cell_data_5 = book_sheet.cell(row, column=5).value
    cell_data_6 = book_sheet.cell(row, column=6).value
    cell_data_7 = book_sheet.cell(row, column=7).value

    # 创建dict {'prov':'XXX','xxx':'xxx'}
    temp_dict = dict()
    temp_dict["prov"] = cell_data_1
    temp_dict["city"] = cell_data_2
    temp_dict["jc"] = cell_data_3
    temp_dict["qc"] = cell_data_4
    temp_dict["dh"] = cell_data_5
    temp_dict["wd"] = cell_data_6
    temp_dict["jd"] = cell_data_7



    # 将dict添加到list
    temp_list.append(temp_dict)

# 将list转换为字符串写入文件
with open('test.json', mode='w', encoding='utf-8', errors='ignore') as f:
    f.write(str(temp_list))


